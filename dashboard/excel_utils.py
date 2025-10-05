from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime
from decimal import Decimal


class ExcelExporter:
    """مساعد لإنشاء ملفات Excel منسقة بشكل جميل"""
    
    # ألوان النظام
    COLORS = {
        'header_bg': '4472C4',  # أزرق داكن
        'header_text': 'FFFFFF',  # أبيض
        'total_bg': 'D9E1F2',  # أزرق فاتح
        'total_text': '1F4788',  # أزرق غامق
        'percentage_bg': 'E2EFDA',  # أخضر فاتح
        'percentage_text': '548235',  # أخضر غامق
        'warning_bg': 'FCE4D6',  # برتقالي فاتح
        'warning_text': 'C65911',  # برتقالي غامق
        'success_bg': 'C6EFCE',  # أخضر فاتح جداً
        'success_text': '006100',  # أخضر غامق
    }
    
    def __init__(self, title="تقرير"):
        self.wb = Workbook()
        self.ws = self.wb.active
        self.title = title
        self.current_row = 1
        
    def create_header(self, headers):
        """إنشاء صف العناوين بتنسيق جميل"""
        for col_num, header in enumerate(headers, 1):
            cell = self.ws.cell(row=self.current_row, column=col_num)
            cell.value = header
            
            # تنسيق العنوان
            cell.font = Font(name='Arial', size=13, bold=True, color=self.COLORS['header_text'])
            cell.fill = PatternFill(start_color=self.COLORS['header_bg'], 
                                   end_color=self.COLORS['header_bg'], 
                                   fill_type='solid')
            cell.alignment = Alignment(
                horizontal='center', 
                vertical='center',
                wrap_text=True,
                text_rotation=0,
                readingOrder=2
            )
            cell.border = self._create_border()
        
        self.ws.row_dimensions[self.current_row].height = 30
        self.current_row += 1
        return self
    
    def add_row(self, values, style='normal', number_formats=None):
        """إضافة صف بيانات
        
        Args:
            values: قائمة القيم
            style: نمط الصف (normal, total, percentage, warning, success)
            number_formats: قائمة من التنسيقات الرقمية لكل عمود (مثل 'currency' أو 'percentage' أو None)
        """
        if number_formats is None:
            number_formats = [None] * len(values)
            
        for col_num, (value, num_format) in enumerate(zip(values, number_formats), 1):
            cell = self.ws.cell(row=self.current_row, column=col_num)
            
            # تحويل القيم المالية من Decimal إلى float
            if isinstance(value, Decimal):
                cell.value = float(value)
            else:
                cell.value = value
            
            # تطبيق التنسيق الرقمي
            if num_format == 'currency':
                cell.number_format = '#,##0.00 "ر.ع"'
            elif num_format == 'percentage':
                cell.number_format = '0.00"%"'
            elif num_format == 'number':
                cell.number_format = '#,##0'
            
            # تنسيق حسب النوع
            if style == 'total':
                cell.font = Font(name='Arial', size=11, bold=True, color=self.COLORS['total_text'])
                cell.fill = PatternFill(start_color=self.COLORS['total_bg'], 
                                       end_color=self.COLORS['total_bg'], 
                                       fill_type='solid')
            elif style == 'percentage':
                cell.font = Font(name='Arial', size=11, color=self.COLORS['percentage_text'])
                cell.fill = PatternFill(start_color=self.COLORS['percentage_bg'], 
                                       end_color=self.COLORS['percentage_bg'], 
                                       fill_type='solid')
            elif style == 'warning':
                cell.font = Font(name='Arial', size=11, color=self.COLORS['warning_text'])
                cell.fill = PatternFill(start_color=self.COLORS['warning_bg'], 
                                       end_color=self.COLORS['warning_bg'], 
                                       fill_type='solid')
            elif style == 'success':
                cell.font = Font(name='Arial', size=11, color=self.COLORS['success_text'])
                cell.fill = PatternFill(start_color=self.COLORS['success_bg'], 
                                       end_color=self.COLORS['success_bg'], 
                                       fill_type='solid')
            else:
                cell.font = Font(name='Arial', size=10)
            
            cell.alignment = Alignment(
                horizontal='center', 
                vertical='center',
                wrap_text=True,
                readingOrder=2
            )
            cell.border = self._create_border()
        
        self.ws.row_dimensions[self.current_row].height = 22
        self.current_row += 1
        return self
    
    def add_total_row(self, label, value, col_span=None, value_type='number'):
        """إضافة صف المجموع
        
        Args:
            label: نص التسمية
            value: القيمة (رقم أو نص)
            col_span: عدد الأعمدة للدمج
            value_type: نوع القيمة ('number', 'currency', 'percentage')
        """
        if col_span:
            # دمج الخلايا للتسمية
            self.ws.merge_cells(start_row=self.current_row, start_column=1,
                              end_row=self.current_row, end_column=col_span-1)
            cell = self.ws.cell(row=self.current_row, column=1)
            cell.value = label
            cell.font = Font(name='Arial', size=12, bold=True, color=self.COLORS['total_text'])
            cell.fill = PatternFill(start_color=self.COLORS['total_bg'], 
                                   end_color=self.COLORS['total_bg'], 
                                   fill_type='solid')
            cell.alignment = Alignment(
                horizontal='center', 
                vertical='center',
                wrap_text=True,
                readingOrder=2
            )
            cell.border = self._create_border()
            
            # خلية القيمة
            value_cell = self.ws.cell(row=self.current_row, column=col_span)
            
            # تحويل Decimal إلى float إذا كانت القيمة رقمية
            if isinstance(value, Decimal):
                value_cell.value = float(value)
            elif isinstance(value, (int, float)):
                value_cell.value = value
            else:
                value_cell.value = value
            
            # تطبيق التنسيق الرقمي
            if value_type == 'currency':
                value_cell.number_format = '#,##0.00 "ر.ع"'
            elif value_type == 'percentage':
                value_cell.number_format = '0.00"%"'
            elif value_type == 'number':
                value_cell.number_format = '#,##0'
                
            value_cell.font = Font(name='Arial', size=12, bold=True, color=self.COLORS['total_text'])
            value_cell.fill = PatternFill(start_color=self.COLORS['total_bg'], 
                                         end_color=self.COLORS['total_bg'], 
                                         fill_type='solid')
            value_cell.alignment = Alignment(
                horizontal='center', 
                vertical='center',
                readingOrder=2
            )
            value_cell.border = self._create_border()
        
        self.ws.row_dimensions[self.current_row].height = 25
        self.current_row += 1
        return self
    
    def add_percentage_row(self, label, percentage, col_span=None):
        """إضافة صف النسبة المئوية
        
        Args:
            label: نص التسمية
            percentage: القيمة الرقمية للنسبة المئوية (مثل 75.5 لـ 75.5%)
            col_span: عدد الأعمدة للدمج
        """
        if col_span:
            # دمج الخلايا للتسمية
            self.ws.merge_cells(start_row=self.current_row, start_column=1,
                              end_row=self.current_row, end_column=col_span-1)
            cell = self.ws.cell(row=self.current_row, column=1)
            cell.value = label
            cell.font = Font(name='Arial', size=11, bold=True, color=self.COLORS['percentage_text'])
            cell.fill = PatternFill(start_color=self.COLORS['percentage_bg'], 
                                   end_color=self.COLORS['percentage_bg'], 
                                   fill_type='solid')
            cell.alignment = Alignment(
                horizontal='center', 
                vertical='center',
                wrap_text=True,
                readingOrder=2
            )
            cell.border = self._create_border()
            
            # خلية النسبة
            percentage_cell = self.ws.cell(row=self.current_row, column=col_span)
            percentage_cell.value = float(percentage) if isinstance(percentage, Decimal) else percentage
            percentage_cell.number_format = '0.00"%"'
            percentage_cell.font = Font(name='Arial', size=11, bold=True, color=self.COLORS['percentage_text'])
            percentage_cell.fill = PatternFill(start_color=self.COLORS['percentage_bg'], 
                                              end_color=self.COLORS['percentage_bg'], 
                                              fill_type='solid')
            percentage_cell.alignment = Alignment(
                horizontal='center', 
                vertical='center',
                readingOrder=2
            )
            percentage_cell.border = self._create_border()
        
        self.ws.row_dimensions[self.current_row].height = 22
        self.current_row += 1
        return self
    
    def add_empty_row(self):
        """إضافة صف فارغ للمسافة"""
        self.current_row += 1
        return self
    
    def set_column_widths(self, widths):
        """تعيين عرض الأعمدة"""
        for col_num, width in enumerate(widths, 1):
            col_letter = get_column_letter(col_num)
            self.ws.column_dimensions[col_letter].width = width
        return self
    
    def add_title(self, title=None, num_columns=None):
        """إضافة عنوان كبير في الأعلى
        
        Args:
            title: نص العنوان
            num_columns: عدد الأعمدة للدمج (إذا لم يحدد، سيتم استخدام max_column أو 7)
        """
        if title:
            self.title = title
        
        # تحديد عدد الأعمدة للدمج
        if num_columns is None:
            num_columns = self.ws.max_column or 7
            
        # دمج الخلايا للعنوان
        self.ws.merge_cells(start_row=1, start_column=1, 
                           end_row=1, end_column=num_columns)
        
        title_cell = self.ws.cell(row=1, column=1)
        title_cell.value = self.title
        title_cell.font = Font(name='Arial', size=18, bold=True, color=self.COLORS['header_text'])
        title_cell.fill = PatternFill(start_color=self.COLORS['header_bg'], 
                                      end_color=self.COLORS['header_bg'], 
                                      fill_type='solid')
        title_cell.alignment = Alignment(
            horizontal='center', 
            vertical='center',
            readingOrder=2
        )
        title_cell.border = self._create_border()
        
        self.ws.row_dimensions[1].height = 40
        self.current_row = 2
        return self
    
    def _create_border(self):
        """إنشاء حدود الخلية"""
        thin_border = Side(style='thin', color='000000')
        return Border(left=thin_border, right=thin_border, 
                     top=thin_border, bottom=thin_border)
    
    def get_response(self, filename=None):
        """الحصول على HttpResponse للتحميل"""
        if not filename:
            filename = f"{self.title}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        self.wb.save(response)
        return response
